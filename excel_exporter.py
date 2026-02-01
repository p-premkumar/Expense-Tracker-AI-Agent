"""
Excel export functionality for expense data
Exports expenses to .xlsx format with formatting, summaries, and charts
"""
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from database import ExpenseDatabase
from config import CURRENCY

class ExcelExporter:
    def __init__(self):
        self.db = ExpenseDatabase()
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def export_all_expenses(self, user_id, filename=None):
        """Export all user expenses to Excel"""
        if not filename:
            filename = f"expenses_{user_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "All Expenses"
        
        # Get all expenses
        expenses = self.db.get_expenses(user_id)
        
        if not expenses:
            ws['A1'] = "No expenses found"
            wb.save(filename)
            return filename
        
        # Headers
        headers = ["ID", "Date", "Category", "Amount", "Description", "Source"]
        self._add_headers(ws, headers)
        
        # Data
        # Use sequential Excel IDs (1..N) so numbering restarts after deletions
        for seq, (exp_id, amount, category, description, date) in enumerate(expenses, start=1):
            row_idx = seq + 1
            date_obj = datetime.fromisoformat(date)
            ws[f'A{row_idx}'] = seq
            ws[f'B{row_idx}'] = date_obj.strftime("%d-%m-%Y %H:%M")
            ws[f'C{row_idx}'] = category
            ws[f'D{row_idx}'] = amount
            ws[f'E{row_idx}'] = description
            ws[f'F{row_idx}'] = "Text"
            
            # Format currency column
            ws[f'D{row_idx}'].number_format = f'"{CURRENCY}"#,##0.00'
            
            # Apply border
            for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                ws[f'{col}{row_idx}'].border = self.thin_border
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 12
        
        # Add summary sheet
        self._add_summary_sheet(wb, user_id, expenses)
        
        # Add monthly breakdown sheet
        self._add_monthly_breakdown(wb, user_id, expenses)
        
        wb.save(filename)
        return filename
    
    def export_monthly_expenses(self, user_id, filename=None):
        """Export expenses for the current month"""
        if not filename:
            filename = f"expenses_monthly_{user_id}_{datetime.now().strftime('%Y%m_%d_%H%M%S')}.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Monthly Expenses"
        
        # Get monthly expenses
        expenses = self.db.get_summary(user_id, days=30)
        all_expenses = self.db.get_expenses(user_id, days=30)
        
        if not all_expenses:
            ws['A1'] = "No expenses found for this month"
            wb.save(filename)
            return filename
        
        # Headers
        headers = ["Category", "Total Amount", "Transaction Count", "Average Per Item"]
        self._add_headers(ws, headers)
        
        # Data
        for row_idx, (category, total, count) in enumerate(expenses, start=2):
            avg = total / count if count > 0 else 0
            ws[f'A{row_idx}'] = category
            ws[f'B{row_idx}'] = total
            ws[f'C{row_idx}'] = count
            ws[f'D{row_idx}'] = avg
            
            # Format currency columns
            ws[f'B{row_idx}'].number_format = f'"{CURRENCY}"#,##0.00'
            ws[f'D{row_idx}'].number_format = f'"{CURRENCY}"#,##0.00'
            
            # Apply border
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row_idx}'].border = self.thin_border
        
        # Add total row
        total_row = len(expenses) + 2
        ws[f'A{total_row}'] = "TOTAL"
        ws[f'A{total_row}'].font = Font(bold=True)
        ws[f'B{total_row}'] = f"=SUM(B2:B{total_row - 1})"
        ws[f'B{total_row}'].font = Font(bold=True)
        ws[f'B{total_row}'].number_format = f'"{CURRENCY}"#,##0.00'
        ws[f'B{total_row}'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        for col in ['A', 'B']:
            ws[f'{col}{total_row}'].border = self.thin_border
        
        # Column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        
        # Add detailed transactions sheet
        self._add_detailed_sheet(wb, all_expenses)
        
        wb.save(filename)
        return filename
    
    def export_custom_period(self, user_id, days, filename=None):
        """Export expenses for a custom period"""
        if not filename:
            filename = f"expenses_{days}days_{user_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"Last {days} Days"
        
        # Get expenses for period
        all_expenses = self.db.get_expenses(user_id, days=days)
        summary = self.db.get_summary(user_id, days=days)
        
        if not all_expenses:
            ws['A1'] = f"No expenses found in the last {days} days"
            wb.save(filename)
            return filename
        
        # Summary section
        ws['A1'] = f"Expense Summary - Last {days} Days"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        
        # Summary headers
        ws['A3'] = "Category"
        ws['B3'] = "Total"
        ws['C3'] = "Count"
        ws['D3'] = "Avg/Item"
        
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}3'].font = Font(bold=True)
            ws[f'{col}3'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            ws[f'{col}3'].border = self.thin_border
        
        # Summary data
        total_sum = 0
        for row_idx, (category, total, count) in enumerate(summary, start=4):
            avg = total / count if count > 0 else 0
            ws[f'A{row_idx}'] = category
            ws[f'B{row_idx}'] = total
            ws[f'C{row_idx}'] = count
            ws[f'D{row_idx}'] = avg
            total_sum += total
            
            ws[f'B{row_idx}'].number_format = f'"{CURRENCY}"#,##0.00'
            ws[f'D{row_idx}'].number_format = f'"{CURRENCY}"#,##0.00'
            
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row_idx}'].border = self.thin_border
        
        # Total row
        total_row = len(summary) + 4
        ws[f'A{total_row}'] = "TOTAL"
        ws[f'A{total_row}'].font = Font(bold=True)
        ws[f'B{total_row}'] = total_sum
        ws[f'B{total_row}'].font = Font(bold=True)
        ws[f'B{total_row}'].number_format = f'"{CURRENCY}"#,##0.00'
        ws[f'B{total_row}'].fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        
        for col in ['A', 'B']:
            ws[f'{col}{total_row}'].border = self.thin_border
        
        # Column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        
        # Add detailed transactions sheet
        self._add_detailed_sheet(wb, all_expenses, sheet_name=f"Details - {days}d")
        
        wb.save(filename)
        return filename
    
    def _add_headers(self, ws, headers):
        """Add formatted headers to worksheet"""
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.thin_border
    
    def _add_summary_sheet(self, wb, user_id, expenses):
        """Add summary sheet to workbook"""
        ws = wb.create_sheet("Summary")
        
        ws['A1'] = "Expense Summary"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        
        # Calculate summary data
        summary_30 = self.db.get_summary(user_id, days=30)
        summary_7 = self.db.get_summary(user_id, days=7)
        
        total_30 = sum(amount for _, amount, _ in summary_30)
        total_7 = sum(amount for _, amount, _ in summary_7)
        
        # Summary stats
        ws['A3'] = "Last 7 Days:"
        ws['B3'] = total_7
        ws['B3'].number_format = f'"{CURRENCY}"#,##0.00'
        ws['B3'].font = Font(bold=True)
        
        ws['A4'] = "Last 30 Days:"
        ws['B4'] = total_30
        ws['B4'].number_format = f'"{CURRENCY}"#,##0.00'
        ws['B4'].font = Font(bold=True)
        
        if total_30 > 0:
            ws['A5'] = "Daily Average (30d):"
            ws['B5'] = total_30 / 30
            ws['B5'].number_format = f'"{CURRENCY}"#,##0.00'
        
        # Category breakdown
        ws['A7'] = "Category Breakdown (30 Days)"
        ws['A7'].font = Font(bold=True, size=11)
        
        ws['A8'] = "Category"
        ws['B8'] = "Amount"
        ws['C8'] = "Count"
        
        for col in ['A', 'B', 'C']:
            ws[f'{col}8'].font = Font(bold=True)
            ws[f'{col}8'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        for row_idx, (category, amount, count) in enumerate(summary_30, start=9):
            ws[f'A{row_idx}'] = category
            ws[f'B{row_idx}'] = amount
            ws[f'C{row_idx}'] = count
            ws[f'B{row_idx}'].number_format = f'"{CURRENCY}"#,##0.00'
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 10
    
    def _add_monthly_breakdown(self, wb, user_id, expenses):
        """Add monthly breakdown sheet"""
        ws = wb.create_sheet("Monthly Breakdown")
        
        ws['A1'] = "Monthly Breakdown"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        
        # Group by month
        monthly_data = {}
        for _, amount, category, _, date in expenses:
            date_obj = datetime.fromisoformat(date)
            month_key = date_obj.strftime("%Y-%m")
            if month_key not in monthly_data:
                monthly_data[month_key] = 0
            monthly_data[month_key] += amount
        
        ws['A3'] = "Month"
        ws['B3'] = "Total Spent"
        
        for col in ['A', 'B']:
            ws[f'{col}3'].font = Font(bold=True)
            ws[f'{col}3'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        for row_idx, (month, total) in enumerate(sorted(monthly_data.items()), start=4):
            ws[f'A{row_idx}'] = month
            ws[f'B{row_idx}'] = total
            ws[f'B{row_idx}'].number_format = f'"{CURRENCY}"#,##0.00'
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
    
    def _add_detailed_sheet(self, wb, expenses, sheet_name="Detailed"):
        """Add detailed transactions sheet"""
        ws = wb.create_sheet(sheet_name)
        
        # Headers
        headers = ["Date", "Category", "Amount", "Description"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.border = self.thin_border
        
        # Data
        for row_idx, (_, amount, category, description, date) in enumerate(expenses, start=2):
            date_obj = datetime.fromisoformat(date)
            ws[f'A{row_idx}'] = date_obj.strftime("%d-%m-%Y %H:%M")
            ws[f'B{row_idx}'] = category
            ws[f'C{row_idx}'] = amount
            ws[f'D{row_idx}'] = description
            
            ws[f'C{row_idx}'].number_format = f'"{CURRENCY}"#,##0.00'
            
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row_idx}'].border = self.thin_border
        
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 35
